import csv
import os
import shutil
import string
from os.path import abspath, join

import openpyxl

from excel_to_csv import excel_to_csv


def test_excel_to_csv():
    root = join(abspath('.'), 'excel_to_csv', 'test_files')
    out = join(root, 'csv_output')

    # find all excel files on path and convert to csv
    excel_files = [file for file in os.listdir(path=root) if file.lower().endswith('.xlsx')]
    excel_to_csv.excel_to_csv(path=root)

    sheet_count = 0

    for file in excel_files:
        wb = openpyxl.load_workbook(filename=join(root, file))
        sheet_count += len(wb.sheetnames)

        # converts worksheet to list of lists and compares to read CSV
        for sheetname in wb.sheetnames:
            csv_filename = f'{file[:-5]}_{wb[sheetname].title}.csv'.replace(' ', '_')
            with open(join(out, csv_filename), 'r') as csv_file:
                reader = csv.reader(csv_file)
                read_csv = [row for row in reader]
                converted_sheet = sheet_converter(wb[sheetname])
                assert read_csv == converted_sheet

    # get resulting files
    csv_list = [file for file in os.listdir(path=out) if file.lower().endswith('.csv')]
    assert len(csv_list) == sheet_count
    shutil.rmtree(path=out)


def test_worksheet_generator():
    expected_wb_name = 'Alphabet.xlsx'
    expected_sheets = [f'{alpha} sheet' for alpha in string.ascii_uppercase]
    root = join(abspath('.'), 'excel_to_csv', 'sample_files')
    for sheet, workbook in excel_to_csv.worksheet_generator(root=root):
        assert sheet.title in expected_sheets
        assert workbook == expected_wb_name


def sheet_converter(worksheet):
    return [[cell.value for cell in row] for row in worksheet.rows]
