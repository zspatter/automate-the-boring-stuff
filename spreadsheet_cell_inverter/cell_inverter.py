#! /usr/bin/env python3
# cell_inverter.py - rotates cells in spreadsheet (transpose)

import os

import openpyxl


def invert_cells(filename):
    """
    Creates a new sheet and rotates the active sheet's cells

    :param str filename: path to source file (.XLSX)
    """
    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb.active

    print(f"Transposing the content of sheet: '{sheet.title}'")
    inverted_sheet = wb.create_sheet(title=f'Inverted {sheet.title}', index=0)

    for row in sheet.rows:
        for cell in row:
            row = cell.row
            col = cell.column

            inverted_sheet.cell(row=col, column=row).value = cell.value

    wb.save(f'result_{filename}')
    print(f"Resulting file saved as 'result_{filename}'")


if __name__ == '__main__':
    os.chdir(os.path.join(os.path.abspath('.'), 'sample_files'))
    invert_cells('example.xlsx')
