import os

import openpyxl

from spreadsheet_cell_inverter import cell_inverter

original_wd = os.path.abspath(os.getcwd())
filename = 'example.xlsx'


def test_invert_cells():
    setup_directory()
    cell_inverter.invert_cells(filename=filename)

    original_wb = openpyxl.load_workbook(filename=filename)
    original_sheet = original_wb.active

    result_wb = openpyxl.load_workbook(filename=f'result_{filename}')
    result_sheet = result_wb.active

    assert original_sheet.max_row == result_sheet.max_column
    assert original_sheet.max_column == result_sheet.max_row

    assert len(original_wb.sheetnames) + 1 == len(result_wb.sheetnames)

    for row in original_sheet.rows:
        for cell in row:
            row = cell.row
            col = cell.column

            assert original_sheet.cell(row=row, column=col).value == \
                   result_sheet.cell(row=col, column=row).value

    revert_directory()


def setup_directory():
    """
    This function is used by all unit tests to temporarily change the
    working directory to access the required test file.
    """
    os.chdir(os.path.abspath(os.path.join('.', 'spreadsheet_cell_inverter', 'test_files')))


def revert_directory():
    """
    This function is used by all unit tests to revert the working directory
    back to to repo root as well as delete the resulting test file. This
    function simply reverts the state of the repo back to what it was prior
    to running any tests.
    """
    os.remove(f'result_{filename}')
    os.chdir(original_wd)
