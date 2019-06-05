#! /usr/bin/env python3

import sys

import openpyxl
from openpyxl.utils.cell import get_column_letter


def insert_blank_rows(index: int, filename: str, offset: int = 1):
    """
    Inserts a blank row at the passed index. If an offset is provided,
    multiple blank rows are inserted.

    :param int index: row number where blank row starts
    :param int offset: number of blank rows
    :param str filename: name of file (including ext)
    """
    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb.active
    rows = tuple(sheet.rows)
    print('Inserting blank rows...')

    # traverses rows in reverse
    for row in rows[::-1]:
        for cell in row:
            row = cell.row
            column = cell.column

            # moves contents and replaces previous cell with blank value
            if index <= row < index + offset:
                sheet.cell(row=row + offset, column=column).value = cell.value
                sheet.cell(row=row, column=column).value = ''
            # moves contents
            elif row >= index + offset:
                sheet.cell(row=row + offset, column=column).value = cell.value

    wb.save('result_' + filename)
    print(f"Resulting file saved as 'result_{filename}'")


def insert_via_move_range(index: int, filename: str, offset: int = 1):
    """
    Achieves the same result as insert_blank_rows() using openpyxl's
    move_range() function.

    The purpose of this exercise was to manipulate the data myself,
    so the original implementation has been kept.

    :param int index: row number where blank row starts
    :param int offset: number of blank rows
    :param str filename: name of file (including ext)
    """
    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb.active
    print('Inserting blank rows...')

    max_column = get_column_letter(sheet.max_column)
    sheet.move_range(f"a{index}:{max_column}{sheet.max_row}", rows=offset, translate=True)

    wb.save('result_' + filename)
    print(f"Resulting file saved as 'result_{filename}'")


def easy_insert(index: int, filename: str, offset: int = 1):
    """
    Achieves the same result as insert_blank_rows() using openpyxl's
    intended insert_rows() function for moving rows.

    The purpose of this exercise was to manipulate the data myself,
    so the original implementation has been kept. However, this is the
    most practical and elegant solution.

    :param int index: row number where blank row starts
    :param int offset: number of blank rows
    :param str filename: name of file (including ext)
    """
    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb.active
    print('Inserting blank rows...')

    sheet.insert_rows(idx=index, amount=offset)
    wb.save('result_' + filename)
    print(f"Resulting file saved as 'result_{filename}'")


if __name__ == '__main__':
    if len(sys.argv) != 4:
        print('Usage: blank_row_inserter.py [index] [offset] [filename]')
    else:
        start = int(sys.argv[1])
        blank_number = int(sys.argv[2])
        file = sys.argv[3]

        insert_blank_rows(index=start, offset=blank_number, filename=file)
