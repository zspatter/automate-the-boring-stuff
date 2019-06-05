from blank_row_inserter import blank_row_inserter
import os
import openpyxl


def test_insert_blank_rows():
    original_wd = os.path.abspath(os.getcwd())
    os.chdir(os.path.abspath(os.path.join('.', 'blank_row_inserter', 'test_files')))

    index, offset, filename = 5, 3, 'test_multiplication_table.xlsx'
    blank_row_inserter.insert_blank_rows(index=index, offset=offset,
                                         filename=filename)

    original = openpyxl.load_workbook('test_multiplication_table.xlsx')
    original_sheet = original.active

    result = openpyxl.load_workbook('result_test_multiplication_table.xlsx')
    result_sheet = result.active

    assert original_sheet.max_row + offset == result_sheet.max_row

    os.remove(f'result_{filename}')
    os.chdir(original_wd)
