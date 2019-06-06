import os

import openpyxl

from row_inserter import row_inserter

index, offset, filename = 5, 3, 'test_multiplication_table.xlsx'
original_wd = os.path.abspath(os.getcwd())


def test_insert_blank_rows():
    setup_directory()
    row_inserter.insert_blank_rows(index=index, offset=offset,
                                   filename=filename)
    compare_sequence()
    revert_directory()


def test_insert_via_move_range():
    setup_directory()
    row_inserter.insert_via_move_range(index=index, offset=offset,
                                       filename=filename)
    compare_sequence()
    revert_directory()


def test_easy_insert():
    setup_directory()

    row_inserter.easy_insert(index=index, offset=offset, filename=filename)
    compare_sequence()
    revert_directory()


def compare_sequence():
    """
    This function is a common test used by all insert row functions. This
    simply compares the contents of the original and resulting file.
    """
    original = openpyxl.load_workbook('test_multiplication_table.xlsx')
    original_sheet = original.active
    result = openpyxl.load_workbook('result_test_multiplication_table.xlsx')
    result_sheet = result.active

    # verifies expected length
    assert original_sheet.max_row + offset == result_sheet.max_row

    # compares unmoved rows
    for row in range(1, index):
        for col in range(1, original_sheet.max_column):
            assert original_sheet.cell(row=row, column=col).value == \
                   result_sheet.cell(row=row, column=col).value

    # compares moved rows
    for row in range(index, original_sheet.max_row + 1):
        for col in range(1, original_sheet.max_column):
            assert original_sheet.cell(row=row, column=col).value == \
                   result_sheet.cell(row=row + offset, column=col).value


def setup_directory():
    """
    This function is used by all unit tests to temporarily change the
    working directory to access the required test file.
    """
    os.chdir(os.path.abspath(os.path.join('.', 'row_inserter', 'test_files')))


def revert_directory():
    """
    This function is used by all unit tests to revert the working directory
    back to to repo root as well as delete the resulting test file. This
    function simply reverts the state of the repo back to what it was prior
    to running any tests.
    """
    os.remove(f'result_{filename}')
    os.chdir(original_wd)
