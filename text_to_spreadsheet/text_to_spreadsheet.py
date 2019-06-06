#! /usr/bin/env python3

import os

import openpyxl
from openpyxl.styles import Font


def text_to_spreadsheet(directory: str = '.',
                        output_file: str = 'text_to_sheet.xlsx'):
    """
    Searches for all text files at the given directory. Each individual
    text file is converted to a column in the output_file spreadsheet.

    :param str directory: path to directory to search
    :param str output_file: name of output file
    """
    path = os.path.abspath(directory)

    wb = openpyxl.Workbook()
    wb.create_sheet(title='Text to Columns', index=0)
    sheet = wb.active
    bold = Font(bold=True)

    # gathers all text files in the given directory
    print(f'Searching for text files...')
    files = [file for file in os.listdir(path) if file.endswith('.txt')]
    column = 1
    print(f'Writing lines of text to columns...')

    # iterates over all the found text files
    for file in files:
        row = 2

        # writes each line of a text file as a cell within a column
        with open(os.path.join(path, file)) as text:
            sheet.cell(row=1, column=column).value = file
            sheet.cell(row=1, column=column).font = bold
            for line in text:
                sheet.cell(row=row, column=column).value = line
                row += 1

        column += 1

    wb.save(filename=output_file)
    print(f"Resulting file saved as '{output_file}'")


if __name__ == '__main__':
    text_to_spreadsheet(os.path.join(os.path.abspath('.'), 'sample_files'))
