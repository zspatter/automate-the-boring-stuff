import os

from text_to_spreadsheet import text_to_spreadsheet
import openpyxl


def text_to_spreadsheet():
    path = os.path.join(os.path.abspath('.'), 'text_to_spreadsheet', 'test_files')
    filename = 'test_text_to_sheet.xlsx'
    text_to_spreadsheet.text_to_spreadsheet(directory=path, output=filename)

    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    files = [file for file in os.listdir(path) if file.lower().endswith('.txt')]
    assert sheet.max_column == len(files)

    column = 1

    for file in files:
        assert sheet.cell(row=1, column=column).value == file
        row = 2

        with open(os.path.join(path, file)) as text:
            for line in text:
                assert line == sheet.cell(row=row, column=column).value
                row += 1

        column += 1

    os.remove(filename)
