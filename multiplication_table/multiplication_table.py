#! /usr/bin/env python3
# multiplication_table.py -

import sys

import openpyxl
from openpyxl.styles import Font, Border, PatternFill, Side, Alignment

# values used for formatting
bold = Font(bold=True)
center = Alignment(horizontal='center', vertical='center')
thick = Border(left=Side(style='medium'),
               right=Side(style='medium'),
               top=Side(style='medium'),
               bottom=Side(style='medium'))
regular = Border(left=Side(style='thin'),
                 right=Side(style='thin'),
                 top=Side(style='thin'),
                 bottom=Side(style='thin'))
grey = PatternFill(start_color='D9D9D9',
                   end_color='D9D9D9',
                   fill_type='solid')


def generate_multiplication_table(n, filename='multiplication_table.xlsx'):
    """
    Generates an NxN multiplication table and saves the workbook with
    the filename.

    :param int n: maximum value
    :param str filename: name of saved file
    """
    print(f'Generating a {n}x{n} multiplication table...')

    # creates and names worksheet appropriately
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f'{n}x{n} multiplication table'

    # freezes pane for both column and row headings
    sheet.freeze_panes = 'B2'

    # creates headers
    for x in range(2, n + 2):
        sheet.cell(row=x, column=1).value = x - 1
        sheet.cell(row=x, column=1).font = bold
        sheet.cell(row=x, column=1).fill = grey
        sheet.cell(row=x, column=1).border = thick
        sheet.cell(row=x, column=1).alignment = center

        sheet.cell(row=1, column=x).value = x - 1
        sheet.cell(row=1, column=x).font = bold
        sheet.cell(row=1, column=x).fill = grey
        sheet.cell(row=1, column=x).border = thick
        sheet.cell(row=1, column=x).alignment = center

    # fills out table values
    for row in range(1, n + 1):
        for column in range(1, n + 1):
            sheet.cell(row=row + 1, column=column + 1).value = row * column
            sheet.cell(row=row + 1, column=column + 1).border = regular

    wb.save(filename)
    print(f"{sheet.title} saved as: '{filename}'")


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Usage: multiplication_table.py [max integer value]')
    else:
        generate_multiplication_table(int(sys.argv[1]))
