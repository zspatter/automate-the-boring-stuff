from multiplication_table import multiplication_table
import openpyxl
import os


def test_generate_multiplication_table():
    file = 'test_multiplication.xlsx'
    multiplication_table.generate_multiplication_table(n=15, filename=file)

    wb = openpyxl.load_workbook('test_multiplication.xlsx')
    sheet = wb.active

    # verify table is square
    assert sheet.max_row == sheet.max_column

    # verify heading values
    for x in range(2, sheet.max_row + 1):
        assert sheet.cell(row=1, column=x).value == x - 1
        assert sheet.cell(row=x, column=1).value == x - 1

    # verify multiplication results
    for row in range(1, sheet.max_row):
        for column in range(1, sheet.max_column):
            assert sheet.cell(row=row + 1, column=column + 1).value == row * column

    os.remove(file)
