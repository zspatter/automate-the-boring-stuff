#! /usr/bin/env python3
# update_produce.py - Corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATES = {'celery': 1.19,
                 'garlic': 3.07,
                 'lemon':  1.27}

for row in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row, column=1).value.lower()
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updatedProduceSales.xlsx')
