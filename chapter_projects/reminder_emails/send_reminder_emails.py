#! /usr/bin/env python3
# send_reminder_emails.py - sends emails based on payment status in spreadsheet

import smtplib
import sys
from os.path import abspath, join

import openpyxl


def get_unpaid_members(worksheet):
    """
    Searches for members who haven't paid their dues for the latest month

    :param  openpyxl.worksheet.worksheet.Worksheet worksheet: worksheet tracking dues payment
    """
    unpaid_members = dict()

    for x in range(2, worksheet.max_row + 1):
        payment = worksheet.cell(row=x, column=worksheet.max_column).value
        if payment != 'paid':
            name = worksheet.cell(row=x, column=1).value.lower()
            email = worksheet.cell(row=x, column=2).value.lower()
            unpaid_members[name] = email

    return unpaid_members


def send_reminders(user, password, month, unpaid_members):
    """
    Sends dues reminder emails to those who haven't paid

    :param str user: email login info
    :param str password: email login info
    :param str month: month that is unpaid
    :param str unpaid_members: name of member
    """
    smtp = smtplib.SMTP('smtp.gmail.com', 587)
    smtp.ehlo()
    smtp.starttls()
    smtp.login(user=user, password=password)

    for name, email in unpaid_members.items():
        body = f'To: {email}' \
            f'\nSubject: {month} Dues Unpaid' \
            f'\nDear {name},' \
            f'\nRecords show that you have not paid does for {month}. ' \
            f'Please make this payment as soon as possible. Thank you!'
        print(f'Sending email to {email}...')
        send_mail_status = smtp.sendmail(from_addr=user,
                                         to_addrs=email,
                                         msg=body)

        if send_mail_status != {}:
            print(f'There was a problem sending email to {email}: {send_mail_status}')

    smtp.quit()


if __name__ == '__main__':
    # open spreadsheet and get latest dues status
    wb = openpyxl.load_workbook(join(abspath('.'), 'duesRecords.xlsx'))

    sheet = wb.active
    last_month = sheet.cell(row=1, column=sheet.max_column).value

    unpaid = get_unpaid_members(worksheet=sheet)
    email_password = sys.argv[1]
    send_reminders(user='cli.disposable@gmail.com',
                   password=email_password,
                   month=last_month,
                   unpaid_members=unpaid)
